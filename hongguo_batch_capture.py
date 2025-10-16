# -*- coding: utf-8 -*-
"""
MuMu 批量点击+截图（ADB 中文输入版 | 稳定中文输入 + 点击搜索+回车 双保险）
- 首次运行：按顺序标定 ①/③/⑤/⑥ → 弹出取框（第7步区域） → ⑧/⑨/⑩；保存 hongguo_config.json
- ② 输入：使用 ADB 直接输入中文（比剪贴板更稳定）
- ③ 搜索：点击你标定的搜索按钮后，再按一次 Enter（双保险）
- ④ 固定等待 5 秒
- Excel：启动时选择，A 列从 A1 开始，遇空停止
- 中止：Ctrl + Alt + Q；或把鼠标移到屏幕左上角（PyAutoGUI FailSafe）
"""
import os, sys, json, time, re, threading, subprocess, importlib, shutil
from pathlib import Path

# ===== 依赖保障 =====
def ensure(pkg, import_name=None):
    name = import_name or pkg
    try:
        return importlib.import_module(name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", pkg])
        importlib.invalidate_caches()
        return importlib.import_module(name)

pyautogui = ensure("pyautogui")
openpyxl   = ensure("openpyxl")
pynput     = ensure("pynput")
from openpyxl import load_workbook
from pynput.keyboard import GlobalHotKeys
from pynput import mouse as pynput_mouse

# 取框 UI（tk）
try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TK = True
except Exception:
    HAS_TK = False
    print("[提示] 未检测到 tkinter，首次标定的取框功能不可用。请安装 tk：python -m pip install tk")

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.12

CONFIG_FILE = Path("hongguo_config.json")

# ===== ADB 中文输入 =====
ADB_PATH = os.environ.get("ADB_PATH") or shutil.which("adb") or "D:\\Program\\MuMuPlayer\
x_main\\adb.exe"
ADB_HOST = os.environ.get("ADB_HOST", "127.0.0.1:16384")

def adb(args, check=False):
    try:
        return subprocess.run([ADB_PATH] + args, capture_output=True, text=True, check=check)
    except Exception:
        return None

def adb_connect():
    # 检查连接状态
    d = adb(["devices"])
    if d and d.stdout:
        # 检查输出中是否包含设备连接信息
        lines = d.stdout.strip().split('
')
        for line in lines:
            if ADB_HOST in line and "device" in line:
                print("  [ADB] 已连接")
                return True
    
    # 尝试连接
    print("  [ADB] 尝试连接...")
    adb(["connect", ADB_HOST])
    time.sleep(1)  # 等待连接
    
    # 再次检查连接状态
    d2 = adb(["devices"])
    if d2 and d2.stdout:
        lines = d2.stdout.strip().split('
')
        for line in lines:
            if ADB_HOST in line and "device" in line:
                print("  [ADB] 连接成功")
                return True
    
    print("  [ADB] 连接失败")
    return False

def adb_tap(x, y):
    adb(["shell", "input", "tap", str(int(x)), str(int(y))])

def adb_key(code):
    adb(["shell", "input", "keyevent", str(code)])

def adb_text(text):
    """使用 ADB 输入文本（支持中文）"""
    if not adb_connect():
        print("  [警告] ADB 未连接，尝试使用剪贴板方式")
        return False
    
    # 清除输入框内容
    adb_tap(x, y)  # 点击输入框确保焦点
    time.sleep(0.1)
    adb_key(67)  # KEYCODE_DEL 删除所有内容
    time.sleep(0.1)
    
    # 使用 ADB 输入文本
    # 对于中文，使用 input text 命令（需要设备支持中文输入法）
    try:
        # 尝试直接输入
        r = adb(["shell", "input", "text", text])
        if r and r.returncode == 0:
            return True
    except:
        pass
    
    print("  [警告] ADB 输入失败，请确保 MuMu 已安装中文输入法")
    return False

# 首次标定顺序：①③⑤⑥ → 取框 → ⑧⑨⑩
STEP_KEYS_ORDER = ["step1","step3","step5","step6","ROI_AFTER_6","step8","step9","step10"]
PROMPTS = {
    "step1":  "① 把鼠标移到【搜索输入框】，左键单击一次（记录坐标）",
    "step3":  "③ 把鼠标移到【搜索按钮/确认】，左键单击一次（记录坐标）",
    "step5":  "⑤ 把鼠标移到第⑤步目标，左键单击一次（记录坐标）",
    "step6":  "⑥ 把鼠标移到第⑥步目标，左键单击一次（记录坐标）",
    "ROI_AFTER_6": "现在弹出【半透明取框】用于第7步截图区域：左键拖拽，松开结束；右键/ESC 取消",
    "step8":  "⑧ 把鼠标移到第⑧步目标，左键单击一次（记录坐标）",
    "step9":  "⑨ 把鼠标移到第⑨步目标，左键单击一次（记录坐标）",
    "step10": "⑩ 把鼠标移到第⑩步目标，左键单击一次（记录坐标）",
}

# ===== 小工具 =====
def safe_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|\r\n\t]+','',(s or '').strip())

def read_actor_names(xlsx_path: str):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    r=1
    while True:
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip()=="":
            break
        yield str(v).strip()
        r+=1

def choose_excel():
    if not HAS_TK:
        print("[错误] 需要 tkinter 才能选择文件。请安装 tk 或把 Excel 路径写入脚本。")
        sys.exit(1)
    root=tk.Tk(); root.withdraw(); root.update()
    p=filedialog.askopenfilename(title="选择演员名称 Excel（A列A1起）",
        filetypes=[("Excel 文件","*.xlsx;*.xlsm;*.xltx;*.xltm")])
    root.destroy(); return p

def countdown(sec=3, tip="将开始执行，请切到 MuMu 窗口…"):
    for i in range(sec,0,-1):
        print(f"{tip} {i}s", end="\r", flush=True); time.sleep(1)
    print("开始执行……".ljust(40))

# ===== 首次标定：取点/取框 =====
def capture_one_click(prompt, timeout=60):
    print("\n" + prompt)
    print("（提示：点击后自动记录；若取消按 Esc 结束程序）")
    pos = {"x": None, "y": None}
    done = threading.Event()
    def on_click(x, y, button, pressed):
        if pressed and button == pynput_mouse.Button.left:
            pos["x"], pos["y"] = int(x), int(y)
            done.set()
            return False
    listener = pynput_mouse.Listener(on_click=on_click)
    listener.start()
    ok = done.wait(timeout); listener.stop()
    if not ok or pos["x"] is None:
        raise RuntimeError("等待点击超时/被取消。")
    print(f"  -> 已记录：({pos['x']}, {pos['y']})")
    return pos["x"], pos["y"]

class ROISelector:
    def __init__(self):
        self.root = tk.Tk()
        self.root.attributes("-topmost", True)
        self.root.attributes("-fullscreen", True)
        try: self.root.attributes("-alpha", 0.25)
        except Exception: pass
        self.root.configure(bg="black")
        self.root.title("拖拽选择截图区域：左键拖拽，松开结束；右键/ESC 取消")
        self.canvas = tk.Canvas(self.root, cursor="cross", bg="black", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.start=None; self.rect=None; self.result=None
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        self.canvas.bind("<ButtonPress-3>", self.on_cancel)
        self.root.bind("<Escape>", self.on_cancel)
    def on_press(self, e):
        self.start=(e.x_root,e.y_root)
        if self.rect: self.canvas.delete(self.rect); self.rect=None
    def on_drag(self, e):
        if not self.start: return
        x1,y1=self.start; x2,y2=e.x_root,e.y_root
        if self.rect: self.canvas.delete(self.rect)
        self.rect=self.canvas.create_rectangle(min(x1,x2),min(y1,y2),max(x1,x2),max(y1,y2),
                                               outline="white", width=2)
    def on_release(self, e):
        if not self.start: return
        x1,y1=self.start; x2,y2=e.x_root,e.y_root
        self.result=(min(x1,x2),min(y1,y2),max(x1,x2),max(y1,y2)); self.root.destroy()
    def on_cancel(self, *_):
        self.result=None; self.root.destroy()
    def show(self):
        self.root.mainloop(); return self.result

def setup_wizard():
    if not HAS_TK:
        print("[错误] 取框需要 tkinter：python -m pip install tk")
        sys.exit(1)
    print("\n=== 首次运行标定：①③⑤⑥ → 取框（第7步） → ⑧⑨⑩ ===")
    coords, region = {}, None
    for key in STEP_KEYS_ORDER:
        if key == "ROI_AFTER_6":
            print("\n" + PROMPTS[key])
            roi = ROISelector().show()
            if not roi: raise RuntimeError("已取消截图区域选择。")
            L,T,R,B = map(int, roi)
            region = {"left":L,"top":T,"right":R,"bottom":B}
            print(f"  -> 截图区域：({L},{T})-({R},{B})")
        else:
            coords[key] = capture_one_click(PROMPTS[key])
    cfg = {"coords": coords, "region": region}
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n标定完成！配置已保存：{CONFIG_FILE.resolve()}")
    return cfg

def load_or_setup_config():
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text("utf-8"))
            need = {"step1","step3","step5","step6","step8","step9","step10"}
            if "coords" in cfg and "region" in cfg and need.issubset(cfg["coords"].keys()):
                return cfg
            print("[提示] 配置不完整，重新标定……")
        except Exception as e:
            print("[提示] 读取配置失败，将重新标定：", e)
    return setup_wizard()

# ===== 中止热键 =====
stop_flag=False
def start_hotkey():
    def on_stop():
        global stop_flag; stop_flag=True; print("\n[中止] 收到热键，准备安全停止…")
    h = GlobalHotKeys({'<ctrl>+<alt>+q': on_stop})
    h.start(); return h

# ===== 行为封装 =====
def click_xy(x, y, clicks=1, wait=0.18):
    pyautogui.moveTo(x, y, duration=0.05)
    pyautogui.click(clicks=clicks)
    time.sleep(wait)

def input_text_via_adb(text: str, x: int, y: int):
    """使用 ADB 输入文本（支持中文）"""
    # 强制使用ADB连接和输入
    if not adb_connect():
        print("  [错误] ADB 连接失败，无法继续")
        return False
    
    # 清除输入框内容
    adb_tap(x, y)  # 点击输入框确保焦点
    time.sleep(0.1)
    adb_key(67)  # KEYCODE_DEL 删除所有内容
    time.sleep(0.1)
    
    # 使用 ADB 输入文本
    try:
        # 对于中文，需要设备安装中文输入法
        r = adb(["shell", "input", "text", text])
        if r and r.returncode == 0:
            print("  [ADB] 文本输入成功")
            return True
        else:
            print("  [错误] ADB 输入失败")
            return False
    except Exception as e:
        print(f"  [错误] ADB 输入异常: {e}")
        return False

def take_region_screenshot(region: dict, save_name: str):
    L,T,R,B = region["left"],region["top"],region["right"],region["bottom"]
    w,h = R-L, B-T
    img = pyautogui.screenshot(region=(L,T,w,h))
    fn = f"{safe_filename(save_name)}.png"
    img.save(fn)
    print(f"[OK] 截图：{fn}")

def run_once(actor: str, cfg: dict):
    c, reg = cfg["coords"], cfg["region"]
    # ① 点击输入框确保焦点
    x,y = c["step1"]; click_xy(x,y, clicks=1, wait=0.12)
    # ② 使用 ADB 输入演员名（支持中文）
    input_text_via_adb(actor, x, y)
    # ③ 点击搜索按钮 + Enter 双保险
    x,y = c["step3"]; click_xy(x,y)
    pyautogui.press("enter"); time.sleep(0.1)
    # ④ 等待 5 秒
    time.sleep(5.0)
    # ⑤
    x,y = c["step5"]; click_xy(x,y)
    # ⑥
    x,y = c["step6"]; click_xy(x,y)
    # ⑦ 截图
    take_region_screenshot(reg, actor)
    # ⑧
    x,y = c["step8"]; click_xy(x,y)
    # ⑨
    x,y = c["step9"]; click_xy(x,y)
    # ⑩
    x,y = c["step10"]; click_xy(x,y)

# ===== 主入口 =====
def main():
    print("MuMu 批量采集（ADB 版 | 支持中文输入）")
    print("中止热键：Ctrl+Alt+Q；或把鼠标移到屏幕左上角。\n")

    excel = choose_excel()
    if not excel or not Path(excel).exists():
        print("[错误] Excel 路径无效"); sys.exit(1)

    cfg = load_or_setup_config()
    # 倒计时，给你切回 MuMu 前台
    countdown(3)

    hot = start_hotkey()
    try:
        for i, actor in enumerate(read_actor_names(excel), 1):
            if stop_flag: break
            print(f"[{i}] 处理：{actor}")
            try:
                run_once(actor, cfg)
            except pyautogui.FailSafeException:
                print("\n[紧急中止] 检测到 FAILSAFE（鼠标到左上角）。"); break
            except Exception as e:
                print(f"[警告] 本条异常：{e}")
            time.sleep(0.3)
    finally:
        try: hot.stop()
        except: pass
    if stop_flag: print("已根据热键请求停止。")
    print("任务结束。")

if __name__ == "__main__":
    main()